"""
Servicio para manejar operaciones relacionadas con ciudades
"""
from typing import List
from openpyxl import load_workbook
from fastapi import HTTPException
from graphmap.domain.model.entities.city import City
from config import settings


class CityService:
    """Servicio para manejar la lógica de negocio relacionada con ciudades"""

    # Caché estático para evitar recargar el Excel en cada request
    _cities_cache: List[City] = None

    def __init__(self, excel_file_path: str = None):
        """
        Inicializa el servicio con la ruta del archivo Excel

        Args:
            excel_file_path: Ruta al archivo Excel con los datos de ciudades.
                           Si es None, usa la configuración del .env
        """
        self.excel_file_path = excel_file_path or settings.DATASET_PATH
    
    def load_cities_from_excel(self) -> List[City]:
        """
        Carga todas las ciudades desde el archivo Excel (con caché)

        Returns:
            Lista de objetos City con todos los datos del Excel

        Raises:
            HTTPException: Si hay error al leer el archivo
        """
        # Retornar desde caché si ya existe
        if CityService._cities_cache is not None:
            return CityService._cities_cache

        try:
            # Leer el archivo Excel con openpyxl para evitar dependencias pesadas en serverless.
            workbook = load_workbook(self.excel_file_path, data_only=True, read_only=True)
            worksheet = workbook.active

            rows = worksheet.iter_rows(values_only=True)
            headers = next(rows, None)
            if not headers:
                return []

            header_index = {str(name): idx for idx, name in enumerate(headers) if name is not None}
            required = [
                "city", "city_ascii", "lat", "lng", "country",
                "iso2", "iso3", "admin_name", "capital", "population", "id"
            ]
            missing = [name for name in required if name not in header_index]
            if missing:
                raise ValueError(f"Missing required columns in dataset: {missing}")

            cities = []

            for row in rows:
                record = {name: row[header_index[name]] for name in required}
                population_raw = record["population"]
                population = int(population_raw) if population_raw is not None else None

                city = City(
                    city=str(record["city"]),
                    city_ascii=str(record["city_ascii"]),
                    lat=float(record["lat"]),
                    lng=float(record["lng"]),
                    country=str(record["country"]),
                    iso2=str(record["iso2"]),
                    iso3=str(record["iso3"]),
                    admin_name=str(record["admin_name"]),
                    capital=str(record["capital"]),
                    population=population,
                    id=int(record["id"])
                )
                cities.append(city)

            workbook.close()

            # Guardar en caché
            CityService._cities_cache = cities
            return cities
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Error loading data from Excel: {str(e)}"
            )
    
    def get_cities_count(self) -> int:
        """
        Obtiene el número total de ciudades
        
        Returns:
            Número total de ciudades en el dataset
        """
        cities = self.load_cities_from_excel()
        return len(cities)
    

    def search_cities(self, query: str) -> List[City]:
        """
        Busca ciudades cuyo nombre contenga la cadena de consulta
        
        Args:
            query: Cadena para buscar en los nombres de las ciudades
            
        Returns:
            Lista de ciudades que coinciden con la consulta
        """
        all_cities = self.load_cities_from_excel()
        query_lower = query.lower()
        cities = [
            city for city in all_cities 
            if query_lower in city.city.lower() or query_lower in city.city_ascii.lower()
        ]
        return cities
